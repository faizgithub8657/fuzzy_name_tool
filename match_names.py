import re
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

# ── CONFIG ──────────────────────────────────────────────────────────────────
EXCEL_FILE = "Book1.xlsx"          # your file name
MAIN_SHEET = "Main sheet"          # sheet with your customer names
TALLY_SHEET = "tally ka data"       # sheet with Tally ledger names
YOUR_COL = "A"                     # column that has YOUR names
RESULT_COL = "B"                   # column where matched Tally name goes
HEADER_ROWS = 1                     # how many header rows to skip
THRESHOLD = 0.5                     # match confidence (0 to 1); raise to be stricter
# ────────────────────────────────────────────────────────────────────────────

# Words so common they'd cause false matches — add more if needed
NOISE_WORDS = {
    "KUMAR", "SINGH", "PATIL", "YADAV", "LAL", "RAM", "DEVI",
    "KUMARI", "PRASAD", "RAO", "SHRI", "SHAH", "DAS"
}


def clean_name(name):
    """Normalize a name and return significant words."""
    text = str(name or "").upper()
    text = re.sub(r"KV[-/][A-Z][-/]\d+[\w-]*", "", text)
    text = re.sub(r"\(.*?\)", "", text)
    text = re.sub(r"[^A-Z\s&]", " ", text)
    words = {w for w in text.split() if len(w) > 1}
    return words - NOISE_WORDS


def match_score(name_a, name_b):
    """Return a similarity score between 0.0 and 1.0."""
    words_a = clean_name(name_a)
    words_b = clean_name(name_b)
    if not words_a or not words_b:
        return 0.0

    shared = len(words_a & words_b)
    union = len(words_a | words_b)
    coverage = shared / len(words_a)
    jaccard = shared / union
    return (coverage + jaccard) / 2


def find_best_match(your_name, tally_names):
    """Return the best matching Tally name and score."""
    best_name = "NOT FOUND"
    best_score = 0.0

    for tally_name in tally_names:
        score = match_score(your_name, tally_name)
        if score >= THRESHOLD and score > best_score:
            best_score = score
            best_name = tally_name

    return best_name, round(best_score, 2)

# ── MAIN ─────────────────────────────────────────────────────────────────────

def load_names(sheet, column, start_row):
    """Read non-empty names from a single column."""
    names = []
    for row in sheet.iter_rows(min_row=start_row, min_col=column, max_col=column, values_only=True):
        value = row[0]
        if not value:
            continue
        text = str(value).strip()
        if text:
            names.append(text)
    return names


def main():
    excel_path = Path(EXCEL_FILE)
    if not excel_path.exists():
        raise FileNotFoundError(f"Excel file not found: {excel_path}")

    workbook = load_workbook(excel_path, data_only=True)
    if MAIN_SHEET not in workbook.sheetnames:
        raise ValueError(f"Sheet not found: {MAIN_SHEET}")
    if TALLY_SHEET not in workbook.sheetnames:
        raise ValueError(f"Sheet not found: {TALLY_SHEET}")

    ws_main = workbook[MAIN_SHEET]
    ws_tally = workbook[TALLY_SHEET]

    col_a = column_index_from_string(YOUR_COL)
    col_b = column_index_from_string(RESULT_COL)
    start_row = HEADER_ROWS + 1

    tally_names = load_names(ws_tally, col_a, start_row)
    print(f"Loaded {len(tally_names)} Tally names.\n")

    matched = 0
    not_found = []

    for row in range(start_row, ws_main.max_row + 1):
        raw_name = ws_main.cell(row=row, column=col_a).value
        if raw_name is None:
            continue

        your_name = str(raw_name).strip()
        if not your_name:
            continue

        result, score = find_best_match(your_name, tally_names)
        ws_main.cell(row=row, column=col_b).value = result

        if result == "NOT FOUND":
            not_found.append(your_name)
            print(f"  ❌  Row {row}: {your_name}")
        else:
            matched += 1
            print(f"  ✅  Row {row}: {your_name}  →  {result}  (score: {score})")

    workbook.save(excel_path)

    print("\n── DONE ─────────────────────────────────────────────────────")
    print(f"Matched   : {matched}")
    print(f"Not found : {len(not_found)}")
    print(f"Saved to  : {EXCEL_FILE}")


if __name__ == "__main__":
    main()


